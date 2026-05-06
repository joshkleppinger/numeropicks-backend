"""
NumeroPicks — FastAPI Backend
Serves prediction, scraping, history, and accuracy data for the web app.
"""

import threading
from datetime import datetime
from typing import Optional

from contextlib import asynccontextmanager
from fastapi import FastAPI, BackgroundTasks, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

from database import (
    init_db, save_prediction_db, load_predictions_db,
    save_scrape_state_db, load_scrape_state_db,
    save_draws_db, load_draws_db,
)
from engine import (
    GAMES, load_draws, save_draws, scrape_game, load_scrape_state,
    analyze_and_predict, save_predictions, compare_predictions,
    compute_accuracy, next_draw_date,
    compare_predictions_with_db,
    brier_score_baseline, evaluate_method_brier,
)

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Startup: init DB, load draws, auto-scrape if needed, schedule 12h refresh."""
    # Init Supabase tables
    init_db()

    # Load draw history — try DB first, fall back to CSV
    for key in GAMES:
        db_rows = load_draws_db(key)
        if db_rows:
            _draw_cache[key] = db_rows
            print(f"[startup] {key}: {len(db_rows):,} draws loaded from DB")
        else:
            csv_rows = load_draws(GAMES[key])
            _draw_cache[key] = csv_rows
            print(f"[startup] {key}: {len(csv_rows):,} draws loaded from CSV")

    def _run_scrape(force=False):
        """Scrape all games and persist to DB + CSV."""
        import time
        if not force:
            time.sleep(1)
        total = sum(len(_draw_cache.get(k, [])) for k in GAMES)
        # Check if any individual game has zero rows - force scrape if so
        any_empty = any(len(_draw_cache.get(k, [])) == 0 for k in GAMES)
        # Check last scrape time from DB
        ls = load_scrape_state_db() or load_scrape_state()
        needs_scrape = (total < 30) or force or any_empty
        if not needs_scrape and ls:
            hours_ago = (datetime.now() - ls).total_seconds() / 3600
            needs_scrape = hours_ago >= 12
        if not needs_scrape:
            print(f"[auto-scrape] {total} rows, data fresh — skipping")
            return
        print(f"[auto-scrape] Scraping all games (total={total}, force={force})...")
        for key in GAMES:
            try:
                rows = _draw_cache.get(key, [])
                added, msg = scrape_game(GAMES[key], rows)
                _draw_cache[key] = rows
                # Persist to DB
                db_added = save_draws_db(key, rows)
                print(f"[auto-scrape] {key}: {msg} (+{db_added} to DB)")
            except Exception as e:
                print(f"[auto-scrape] {key} ERROR: {e}")
        save_scrape_state_db()
        print(f"[auto-scrape] Done. Total: {sum(len(_draw_cache.get(k,[])) for k in GAMES)}")

    def _scheduler():
        """Run a scrape every 12 hours."""
        import time
        _run_scrape(force=False)
        while True:
            time.sleep(12 * 3600)
            _run_scrape(force=True)

    import threading
    threading.Thread(target=_scheduler, daemon=False).start()
    yield  # app runs here

app = FastAPI(title="NumeroPicks API", version="1.0.0", lifespan=lifespan)

# ── CORS: allow the React frontend (any origin in dev, locked down in prod) ───
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,      # must be False when allow_origins=["*"]
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── In-memory draw cache (loaded once on startup, updated after scrapes) ──────
_draw_cache: dict[str, list] = {}
_analyze_lock = threading.Lock()

# ── Progress tracking for long-running analysis ───────────────────────────────
_analysis_progress: dict[str, dict] = {}
_job_results: dict[str, dict] = {}   # job_id -> result or error


def _get_draws(game_key: str) -> list:
    if game_key not in _draw_cache:
        _draw_cache[game_key] = load_draws(GAMES[game_key])
    return _draw_cache[game_key]


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/")
def root():
    return {"status": "NumeroPicks API is running 🎱"}


@app.get("/games")
def list_games():
    """Return metadata for all supported games."""
    return {
        key: {
            "key":          g["key"],
            "display_name": g["display_name"],
            "white_max":    g["white_max"],
            "special_max":  g["special_max"],
            "special_name": g["special_name"],
            "next_draw":    next_draw_date(g),
            "row_count":    len(_get_draws(key)),
        }
        for key, g in GAMES.items()
    }


@app.get("/history/{game_key}")
def get_history(game_key: str, limit: int = 20):
    """Return the most recent draws for a game."""
    if game_key not in GAMES:
        raise HTTPException(status_code=404, detail=f"Unknown game: {game_key}")
    rows = _get_draws(game_key)
    return {
        "game":       game_key,
        "total_rows": len(rows),
        "draws":      rows[-limit:][::-1],   # most recent first
    }


@app.post("/scrape/{game_key}")
def scrape(game_key: str, background_tasks: BackgroundTasks):
    """Kick off a background scrape for one game."""
    if game_key not in GAMES:
        raise HTTPException(status_code=404, detail=f"Unknown game: {game_key}")

    def _run():
        rows = _get_draws(game_key)
        added, msg = scrape_game(GAMES[game_key], rows)
        _draw_cache[game_key] = rows
        print(f"[scrape/{game_key}] {msg}")

    background_tasks.add_task(_run)
    return {"status": "scrape started", "game": game_key}


@app.post("/scrape-all")
def scrape_all(background_tasks: BackgroundTasks):
    """Scrape all three games in the background."""
    def _run():
        for key, game in GAMES.items():
            rows = _get_draws(key)
            added, msg = scrape_game(game, rows)
            _draw_cache[key] = rows
            print(f"[scrape-all/{key}] {msg}")

    background_tasks.add_task(_run)
    return {"status": "scraping all games in background"}


@app.get("/scrape-status")
def scrape_status():
    """Return the last scrape timestamp and staleness flag."""
    ls = load_scrape_state_db() or load_scrape_state()
    if ls is None:
        return {"last_scrape": None, "stale": True, "days_ago": None}
    hours_ago = (datetime.now() - ls).total_seconds() / 3600
    days_ago  = int(hours_ago / 24)
    return {
        "last_scrape": ls.isoformat(),
        "hours_ago":   round(hours_ago, 1),
        "days_ago":    days_ago,
        "stale":       hours_ago >= 12,
    }


@app.get("/predict/{game_key}")
def predict(game_key: str, save: bool = True):
    """
    Run the full 7-method analysis and return 5 predicted tickets.
    Waits up to 90s for auto-scrape to populate data if needed.
    """
    if game_key not in GAMES:
        raise HTTPException(status_code=404, detail=f"Unknown game: {game_key}")

    game = GAMES[game_key]
    rows = _get_draws(game_key)
    if not rows:
        raise HTTPException(status_code=400,
                            detail=f"No draw history for {game_key}. Run /scrape-all first.")

    with _analyze_lock:
        tickets = analyze_and_predict(rows, game)

    nd = next_draw_date(game)

    # Convert numpy integers to plain Python ints for JSON serialization
    clean_tickets = [
        {
            "balls":   [int(b) for b in t["balls"]],
            "special": int(t["special"]),
        }
        for t in tickets
    ]

    if save and tickets:
        save_predictions(game, tickets, nd)
        # Also persist to Supabase so predictions survive restarts
        today = datetime.now().strftime("%Y-%m-%d")
        for t in clean_tickets:
            save_prediction_db(game_key, t["balls"], t["special"], today, nd)

    # Compute Brier Score for this game's ensemble
    brier_info = None
    try:
        bs_recent = evaluate_method_brier(
            {n: 1.0 for n in range(1, game["white_max"]+1)},  # placeholder
            rows, range(1, game["white_max"]+1), 50)
        bs_base   = brier_score_baseline(game["white_count"], game["white_max"])
        brier_info = {
            "recent_brier":   round(bs_base, 5),   # baseline for reference
            "baseline_brier": round(bs_base, 5),
            "improvement_pct": 0.0,
        }
    except Exception:
        pass

    return {
        "game":         game_key,
        "next_draw":    nd,
        "tickets":      clean_tickets,
        "special_name": game["special_name"],
        "brier":        brier_info,
    }


@app.get("/accuracy/{game_key}")
def accuracy(game_key: str):
    """Return prediction accuracy stats and recent evaluated rounds."""
    if game_key not in GAMES:
        raise HTTPException(status_code=404, detail=f"Unknown game: {game_key}")

    game = GAMES[game_key]
    rows = _get_draws(game_key)
    # Load predictions from DB (persistent) + CSV fallback
    db_preds = load_predictions_db(game_key)
    data = compare_predictions_with_db(game, rows, db_preds) if db_preds else compare_predictions(game, rows)

    evaluated = data["evaluated"]
    pending   = data["pending"]

    summary = None
    if evaluated:
        total  = len(evaluated)
        avg_w  = sum(r["white_matches"] for r in evaluated) / total
        sp_hit = sum(1 for r in evaluated if r["sp_match"]) / total * 100
        best_w = max(r["white_matches"] for r in evaluated)
        last   = evaluated[-1]
        summary = {
            "total_rounds":      total,
            "avg_white_matches": round(avg_w, 2),
            "special_hit_rate":  round(sp_hit, 1),
            "best_white_match":  best_w,
            "last_score":        round(last["score"] * 100),
            "last_draw_date":    last["target_draw_date"],
        }

    return {
        "game":      game_key,
        "summary":   summary,
        "evaluated": evaluated[-10:],    # last 10 rounds
        "pending":   pending,
    }


@app.post("/predict-async/{game_key}")
def predict_async(game_key: str, background_tasks: BackgroundTasks):
    """Start a prediction job in the background. Returns a job_id to poll."""
    import uuid
    if game_key not in GAMES:
        raise HTTPException(status_code=404, detail=f"Unknown game: {game_key}")
    rows = _get_draws(game_key)
    if not rows:
        raise HTTPException(status_code=400,
            detail=f"No draw history for {game_key}. Scrape first.")
    job_id = str(uuid.uuid4())[:8]
    _job_results[job_id] = {"status": "running"}

    def _run():
        try:
            game    = GAMES[game_key]
            tickets = analyze_and_predict(rows, game)
            nd      = next_draw_date(game)
            if tickets:
                save_predictions(game, tickets, nd)
            _job_results[job_id] = {
                "status":       "done",
                "game":         game_key,
                "next_draw":    nd,
                "tickets":      tickets,
                "special_name": game["special_name"],
            }
        except Exception as e:
            _job_results[job_id] = {"status": "error", "detail": str(e)}

    background_tasks.add_task(_run)
    return {"job_id": job_id, "status": "running"}


@app.get("/predict-result/{job_id}")
def predict_result(job_id: str):
    """Poll for the result of a predict-async job."""
    result = _job_results.get(job_id)
    if result is None:
        raise HTTPException(status_code=404, detail="Job not found")
    return result


@app.get("/next-draw/{game_key}")
def get_next_draw(game_key: str):
    """Return the next draw date with a friendly formatted string."""
    if game_key not in GAMES:
        raise HTTPException(status_code=404, detail=f"Unknown game: {game_key}")
    game = GAMES[game_key]
    nd   = next_draw_date(game)
    # Build ordinal
    from datetime import datetime as dt
    d       = dt.strptime(nd, "%a, %b %d, %Y")
    day_num = d.day
    sfx     = "th" if 11 <= day_num <= 13 else {1:"st",2:"nd",3:"rd"}.get(day_num%10,"th")
    friendly = f"These numbers are for the {d.strftime('%A')}, {d.strftime('%B')} {day_num}{sfx} drawing"
    return {"date_str": nd, "friendly": friendly}


# ══════════════════════════════════════════════════════════════════════════════
#  STARTUP
# ══════════════════════════════════════════════════════════════════════════════

@app.on_event("startup")
def startup():
    """
    Load draw histories from persistent disk into cache.
    On paid Render tier, data persists between restarts so we rarely need to scrape.
    Only auto-scrapes if data is completely missing.
    """
    total = 0
    for key in GAMES:
        rows = load_draws(GAMES[key])
        _draw_cache[key] = rows
        total += len(rows)
        print(f"[startup] {key}: {len(rows):,} draws loaded")

    print(f"[startup] Total: {total:,} draws across all games")

    if total < 30:
        # First run ever — scrape everything in background
        print("[startup] No data found — running initial scrape in background...")
        def _initial_scrape():
            import time
            time.sleep(2)
            for key in GAMES:
                try:
                    rows = _draw_cache.get(key, [])
                    added, msg = scrape_game(GAMES[key], rows)
                    _draw_cache[key] = rows
                    print(f"[initial-scrape] {key}: {msg}")
                except Exception as e:
                    print(f"[initial-scrape] {key} ERROR: {e}")
        import threading
        threading.Thread(target=_initial_scrape, daemon=False).start()
    else:
        # Data exists — check if it's stale (>3 days) and update in background
        def _refresh_check():
            import time
            time.sleep(5)
            ls = load_scrape_state()
            from datetime import datetime
            if ls is None or (datetime.now() - ls).days >= 3:
                print("[startup] Data stale — refreshing in background...")
                for key in GAMES:
                    try:
                        rows = _draw_cache.get(key, [])
                        added, msg = scrape_game(GAMES[key], rows)
                        _draw_cache[key] = rows
                        if added:
                            print(f"[refresh] {key}: {msg}")
                    except Exception as e:
                        print(f"[refresh] {key} ERROR: {e}")
            else:
                print(f"[startup] Data fresh (last scraped: {ls.strftime('%Y-%m-%d')}), skipping refresh")
        import threading
        threading.Thread(target=_refresh_check, daemon=True).start()


# ══════════════════════════════════════════════════════════════════════════════
#  SPREADSHEET DOWNLOADS
# ══════════════════════════════════════════════════════════════════════════════

from fastapi.responses import StreamingResponse, HTMLResponse
import csv as csv_module
import io


@app.get("/download/{game_key}/csv")
def download_csv(game_key: str):
    """Download all historical draws for a game as a CSV file."""
    if game_key not in GAMES:
        raise HTTPException(status_code=404, detail=f"Unknown game: {game_key}")

    game = GAMES[game_key]
    rows = _get_draws(game_key)
    if not rows:
        raise HTTPException(status_code=404,
                            detail="No data yet. Run a scrape first.")

    n          = game["white_count"]
    sn         = game["special_name"]
    fieldnames = ["date"] + [f"ball_{i}" for i in range(1, n + 1)] + [sn.lower()]

    buf = io.StringIO()
    writer = csv_module.DictWriter(buf, fieldnames=fieldnames)
    writer.writeheader()
    for r in rows:
        row = {"date": r["date"], sn.lower(): r["special"]}
        for i, b in enumerate(r["balls"], 1):
            row[f"ball_{i}"] = b
        writer.writerow(row)

    buf.seek(0)
    filename = f"{game['display_name'].replace(' ', '_')}_draws.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/download/{game_key}/xlsx")
def download_xlsx(game_key: str):
    """Download all historical draws for a game as an Excel (.xlsx) file."""
    if game_key not in GAMES:
        raise HTTPException(status_code=404, detail=f"Unknown game: {game_key}")

    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side)
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Add it to requirements.txt."
        )

    game = GAMES[game_key]
    rows = _get_draws(game_key)
    if not rows:
        raise HTTPException(status_code=404, detail="No data yet.")

    n   = game["white_count"]
    sn  = game["special_name"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = game["display_name"]

    # ── Styling ───────────────────────────────────────────────────────────────
    header_fill   = PatternFill("solid", fgColor="1E293B")   # dark slate
    header_font   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    special_fill  = PatternFill("solid", fgColor="EF4444")   # red
    special_font  = Font(bold=True, color="FFFFFF", name="Calibri")
    alt_fill      = PatternFill("solid", fgColor="F1F5F9")   # light grey
    center        = Alignment(horizontal="center")
    thin_border   = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["Date"] + [f"Ball {i}" for i in range(1, n + 1)] + [sn]
    for col, hdr in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=hdr)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin_border

    # ── Data rows (newest first) ───────────────────────────────────────────────
    for row_idx, r in enumerate(reversed(rows), 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        # Date
        c           = ws.cell(row=row_idx, column=1, value=r["date"])
        c.border    = thin_border
        if fill: c.fill = fill

        # White balls
        for col, b in enumerate(r["balls"], 2):
            c           = ws.cell(row=row_idx, column=col, value=b)
            c.alignment = center
            c.border    = thin_border
            if fill: c.fill = fill

        # Special ball (red)
        c           = ws.cell(row=row_idx, column=n + 2, value=r["special"])
        c.font      = special_font
        c.fill      = special_fill
        c.alignment = center
        c.border    = thin_border

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    for col in range(2, n + 3):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = 9

    # ── Freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{game['display_name'].replace(' ', '_')}_draws.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/downloads", response_class=HTMLResponse)
def downloads_page():
    """A styled HTML page with download buttons for all games."""
    game_blocks = ""
    for key, game in GAMES.items():
        rows      = _get_draws(key)
        row_count = f"{len(rows):,}"
        latest    = rows[-1]["date"] if rows else "No data yet"
        sn        = game["special_name"]
        special_text = f"Special ball: <strong>{sn}</strong>" if sn else "<strong>No special ball</strong>"
        game_blocks += f"""
        <div class="game-card">
          <h2>{game['display_name']}</h2>
          <p class="meta">
            <strong>{row_count}</strong> draws &nbsp;·&nbsp;
            Latest: <strong>{latest}</strong> &nbsp;·&nbsp;
            {special_text}
          </p>
          <div class="btn-row">
            <a href="/download/{key}/csv"  class="btn btn-csv">
              ⬇ Download CSV
            </a>
            <a href="/download/{key}/xlsx" class="btn btn-xlsx">
              ⬇ Download Excel (.xlsx)
            </a>
          </div>
        </div>
        """

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>NumeroPicks — Download Historical Data</title>
  <style>
    *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      background: #0a0e1a; color: #e2e8f0; min-height: 100vh; padding: 2rem;
    }}
    header {{
      text-align: center; margin-bottom: 2.5rem;
    }}
    header h1 {{
      font-size: 2.4rem; font-weight: 900;
      color: #ef4444; letter-spacing: 0.08em; font-family: "Courier New", monospace;
    }}
    header p {{ color: #94a3b8; margin-top: 0.4rem; font-size: 1rem; }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
      gap: 1.5rem; max-width: 1000px; margin: 0 auto;
    }}
    .game-card {{
      background: #111827; border-radius: 1rem; padding: 1.6rem;
      border: 1px solid #1e293b;
    }}
    .game-card h2 {{
      font-size: 1.3rem; font-weight: 700; margin-bottom: 0.5rem;
      color: #f1f5f9;
    }}
    .meta {{ color: #94a3b8; font-size: 0.88rem; margin-bottom: 1.2rem; }}
    .btn-row {{ display: flex; gap: 0.75rem; flex-wrap: wrap; }}
    .btn {{
      display: inline-block; padding: 0.6rem 1.2rem;
      border-radius: 999px; font-weight: 600; font-size: 0.9rem;
      text-decoration: none; transition: opacity .15s;
    }}
    .btn:hover {{ opacity: 0.85; }}
    .btn-csv  {{ background: #1e40af; color: #fff; }}
    .btn-xlsx {{ background: #166534; color: #fff; }}
    footer {{
      text-align: center; margin-top: 3rem;
      color: #475569; font-size: 0.82rem;
    }}
  </style>
</head>
<body>
  <header>
    <h1>🎱 NUMEROPICKS</h1>
    <p>Download complete historical draw data for all supported games</p>
  </header>
  <div class="grid">
    {game_blocks}
  </div>
  <footer>
    <p>Data sourced from lottery.net &nbsp;·&nbsp; numeropicks.com</p>
  </footer>
</body>
</html>"""


# ── Logo endpoint — serves red_ball_logo.png from the data directory ──────────
from fastapi.responses import FileResponse
import mimetypes

@app.get("/logo")
def serve_logo():
    """Serve red_ball_logo.png from the Numero data folder."""
    for fname in ["red_ball_logo.png", "red_ball_logo.jpg"]:
        path = DATA_DIR / fname
        if path.exists():
            mt = mimetypes.guess_type(str(path))[0] or "image/png"
            return FileResponse(str(path), media_type=mt)
    # Return 404 — frontend will fall back to the inline SVG ball
    raise HTTPException(status_code=404, detail="Logo not found")
